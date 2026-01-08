import argparse
import re
import os
import sys
from openpyxl import load_workbook # type: ignore
import pandas as pd # type: ignore
from unidecode import unidecode

parser = argparse.ArgumentParser(description="Update foosball ranking Excel sheet with tournament results.")
parser.add_argument("--results", required=True, help="Path to the .txt file with results.")
parser.add_argument("--excel", required=True, help="Path to the existing Excel file.")
parser.add_argument("--tournament", required=True, help="Name of the tournament.")
parser.add_argument("--level", required=True, help="Tournament level as defined in the Points Lookup sheet.")
parser.add_argument("--category", required=True, help="Category sheet name.")
parser.add_argument("--output", required=True, help="Name of output file.")

args = parser.parse_args()

# === Assign Inputs ===
results_txt = args.results
excel_file = args.excel
output_file = args.output

tournament_name = args.tournament
tournament_level = args.level
category = args.category

# === STEP 1: Read Results File ===
with open(results_txt, 'r', encoding='utf-8') as f:
    lines = [line.strip() for line in f if line.strip()]

placements = []
for line in lines:
    match = re.match(r'^(\d+)\.(.*)', line)
    if match:
        place = int(match.group(1))
        players = match.group(2).split('|')
    else:
        # No explicit ranking: use last place
        place = placements[-1][0]
        players = line.split('|')
    players = [p.strip() for p in players]
    placements.append((place, players))

print(placements) # DEBUG

# === STEP 2: Load Excel File ===
wb = load_workbook(excel_file)
points_sheet = wb["Points Lookup"]
category_sheet = wb[category]

# Parse Points Lookup
points_df = pd.DataFrame(points_sheet.values)
points_df.columns = points_df.iloc[0]
points_df = points_df[1:]

level_column_name = "Level" if "Level" in points_df.columns else list(points_df.columns)[2]
points_df = points_df[points_df[level_column_name] == tournament_level]

if points_df.empty:
    raise ValueError(f"Tournament level '{tournament_level}' not found in column '{level_column_name}'.")

level_row = points_df.iloc[0, 3:13]
places = [1, 2, 3, 4, 5, 9, 17, 33, 65, 129]
points_mapping = {}

for col in range(10):
    try:
        points = int(level_row.iloc[col]) if pd.notna(level_row.iloc[col]) else 0
        points_mapping[places[col]] = points
    except ValueError:
        raise ValueError(f"Non-numeric value '{level_row.iloc[col]}' in points column.")


def bucket_place(place: int) -> int:
    """Map an arbitrary placement to the nearest FRFM points bucket.

    Example: 6/7/8 -> 5, 10..16 -> 9, etc.
    """
    # find the largest bucket <= place
    bucket = places[0]
    for p in places:
        if p <= place:
            bucket = p
        else:
            break
    return bucket

print(points_mapping)

for (place, name) in placements:
    bp = bucket_place(place)
    print(name[0], points_mapping[bp])

category_df = pd.DataFrame(category_sheet.values)
category_df.columns = category_df.iloc[0]

print(category_df.columns.size)
print(category_sheet.max_column)
print(category_sheet.max_row)
print(category_sheet.cell(row=81, column=1).value)
print(category_sheet.cell(row=82, column=1).value)
print(category_sheet.cell(row=81, column=3).value)

# Fill new column with zeroes
new_tournament_column = category_sheet.max_column + 1
first_empty_row = 0

for row in range(7, category_sheet.max_row + 1):
    if category_sheet.cell(row=row, column=1).value == None:
        first_empty_row = row
        break
    category_sheet.cell(row=row, column=new_tournament_column).value = 0

category_sheet.cell(row=1, column=new_tournament_column).value = tournament_name
category_sheet.cell(row=2, column=new_tournament_column).value = 'Phoenix Foosball Cluj'
category_sheet.cell(row=6, column=new_tournament_column).value = tournament_level

print(first_empty_row)

for (place, name) in placements:
    bp = bucket_place(place)
    found = False
    # Loop through column B (column index 2)
    for row in range(1, first_empty_row):
        #print(row)
        cell_value = category_sheet.cell(row=row, column=2).value
        parts = name[0].strip().lower().split() # split the name of the player into 2 or more parts
        if len(parts) > 2 :
            #print(str(cell_value).strip().lower())
            #print(name[0].strip().lower())
            if cell_value and str(cell_value).strip().lower() == name[0].strip().lower():
                category_sheet.cell(row=row, column=new_tournament_column).value = points_mapping[bp]
                found = True
                print(f"Player '{name[0]}' found on row {row}")
        else:
            first_last = " ".join(parts)
            last_first = " ".join(parts[::-1])
            if (cell_value and unidecode(str(cell_value).strip().lower()) == unidecode(first_last)) or (cell_value and unidecode(str(cell_value).strip().lower()) == unidecode(last_first)) :
                category_sheet.cell(row=row, column=new_tournament_column).value = points_mapping[bp]
                found = True
                print(f"Player '{name[0]}' found on row {row}")
    if not found:
        print(f"Player {name[0]} wasn't found in excel. Adding as a new row...")
        category_sheet.cell(row=first_empty_row, column=1).value = first_empty_row - 5
        category_sheet.cell(row=first_empty_row, column=2).value = unidecode(name[0])
        category_sheet.cell(row=first_empty_row, column=new_tournament_column).value = points_mapping[bp]

        # Fill empty cells in new row
        for col in range(9, new_tournament_column):
            category_sheet.cell(row=first_empty_row, column=col).value = 0

        first_empty_row += 1

wb.save(output_file)



#print(f"Player '{player_name}' not found in column B.")


