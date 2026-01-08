#!/usr/bin/env python3

import argparse
import json
import os
import re
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import requests
from openpyxl import load_workbook  # type: ignore
from unidecode import unidecode


PLACE_RE = re.compile(r"^(\d+)\.(.*)$")

TOURNAMENT_URL_RE = re.compile(
    r"https?://live\.kickertool\.de/(?P<club>[^/]+)/tournaments/(?P<tid>[^/]+)/standings"
)


def _norm(name: str) -> str:
    return " ".join(unidecode(name).lower().split())


def _swap_two_tokens(norm_name: str) -> Optional[str]:
    parts = norm_name.split()
    if len(parts) != 2:
        return None
    return f"{parts[1]} {parts[0]}"


def _pretty_token(tok: str) -> str:
    parts = tok.split("-")
    out = []
    for p in parts:
        if not p:
            out.append(p)
        else:
            out.append(p[:1].upper() + p[1:].lower())
    return "-".join(out)


def pretty_name(name: str) -> str:
    name = " ".join(name.split())
    letters_only = re.sub(r"[^A-Za-z]", "", unidecode(name))
    if letters_only and letters_only.isupper():
        return " ".join(_pretty_token(t) for t in name.split())

    return " ".join(_pretty_token(t) for t in name.split())


def _is_tony_spredeman(name: str) -> bool:
    return _norm(name) == "tony spredeman"


@dataclass(frozen=True)
class Entry:
    place: int
    players: List[str]  # singles: len=1, doubles: len=2


def parse_existing_allowlists(root: str) -> Tuple[Set[str], Set[str]]:
    women: Set[str] = set()
    juniors: Set[str] = set()

    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn == "womens.txt":
                path = os.path.join(dirpath, fn)
                women |= _names_from_txt(path)
            elif fn in ("junior.txt", "juniors.txt"):
                path = os.path.join(dirpath, fn)
                juniors |= _names_from_txt(path)

    return women, juniors


def load_women_allowlist_from_excel(excel_path: str) -> Set[str]:
    women: Set[str] = set()

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)

    for sheet_name in wb.sheetnames:
        if "women" not in sheet_name.lower():
            continue
        ws = wb[sheet_name]

        # Try to find a header column with 'name' or 'player'
        name_col: Optional[int] = None
        try:
            header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for idx, v in enumerate(header_row, start=1):
                if not isinstance(v, str):
                    continue
                h = v.strip().lower()
                if h in ("name", "player", "player name") or "name" in h:
                    name_col = idx
                    break
        except Exception:
            name_col = None

        # Fallback to FRFM format: names are in column B starting row 7
        if name_col is None:
            name_col = 2
            start_row = 7
        else:
            # If we found a header, start after it
            start_row = 2

        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=name_col).value
            if v is None:
                # FRFM sheets tend to have a contiguous block; stop when empty
                if r >= start_row + 3:
                    break
                continue
            if not isinstance(v, str):
                continue
            name = v.strip()
            if not name:
                continue
            n = _norm(name)
            women.add(n)
            swapped = _swap_two_tokens(n)
            if swapped:
                women.add(swapped)

    return women


def _names_from_txt(path: str) -> Set[str]:
    out: Set[str] = set()
    try:
        with open(path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line:
                    continue
                m = PLACE_RE.match(line)
                if m:
                    rest = m.group(2).strip()
                else:
                    rest = line
                for p in rest.split("|"):
                    p = p.strip()
                    if p:
                        out.add(_norm(p))
                        swapped = _swap_two_tokens(_norm(p))
                        if swapped:
                            out.add(swapped)
    except OSError:
        return set()
    return out


def load_next_data(url: str, timeout_s: int = 30) -> dict:
    raise RuntimeError("load_next_data() is deprecated; use Kickertool tournament API")


def fetch_tournament_json(tournament_id: str, timeout_s: int = 30) -> dict:
    api_url = f"https://live.kickertool.de/api/table_soccer/tournaments/{tournament_id}.json"
    r = requests.get(
        api_url,
        timeout=timeout_s,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json,text/plain,*/*",
        },
    )
    r.raise_for_status()
    return r.json()


def extract_final_standings(tj: dict) -> List[Entry]:
    """Extract final standings as (place, players) from tournament JSON."""

    # Build participant_id -> players[] map from qualifying participants
    participants: Dict[str, List[str]] = {}
    qualifying = tj.get("qualifying") or []
    if qualifying and isinstance(qualifying, list) and isinstance(qualifying[0], dict):
        plist = qualifying[0].get("participants") or []
        if isinstance(plist, list):
            for p in plist:
                if not isinstance(p, dict):
                    continue
                pid = p.get("_id")
                if not isinstance(pid, str):
                    continue

                players: List[str] = []
                pplayers = p.get("players")
                if isinstance(pplayers, list) and pplayers:
                    for pp in pplayers:
                        if isinstance(pp, dict):
                            n = pp.get("name") or pp.get("displayName")
                            if isinstance(n, str) and n.strip():
                                players.append(pretty_name(n.strip()))
                        elif isinstance(pp, str) and pp.strip():
                            players.append(pretty_name(pp.strip()))

                if not players:
                    # Fallback: participant may just have a name
                    n = p.get("name")
                    if isinstance(n, str) and n.strip():
                        players = [pretty_name(n.strip())]

                if players:
                    participants[pid] = players

    # Use eliminations[0].standings where stats.finalResult=True
    eliminations = tj.get("eliminations") or []
    if not (isinstance(eliminations, list) and eliminations):
        raise RuntimeError("Tournament JSON has no eliminations")

    elim0 = eliminations[0]
    if not isinstance(elim0, dict):
        raise RuntimeError("Tournament JSON eliminations[0] malformed")

    standings = elim0.get("standings") or []
    if not isinstance(standings, list) or not standings:
        raise RuntimeError("Tournament JSON has no elimination standings")

    entries: List[Entry] = []
    for row in standings:
        if not isinstance(row, dict):
            continue
        stats = row.get("stats")
        if not isinstance(stats, dict):
            continue

        if stats.get("finalResult") is not True:
            continue

        place_val = stats.get("place")
        try:
            place = int(place_val)
        except Exception:
            continue

        rid = row.get("_id")
        players = participants.get(rid) if isinstance(rid, str) else None
        if not players:
            # fallback to row name
            n = row.get("name")
            if isinstance(n, str) and n.strip():
                players = [pretty_name(n.strip())]

        if players:
            entries.append(Entry(place=place, players=players))

    entries.sort(key=lambda e: (e.place, "|".join(e.players)))
    if not entries:
        raise RuntimeError("Could not extract final standings entries")
    return entries


def remove_tony(entries: Sequence[Entry]) -> List[Entry]:
    # Remove Tony Spredeman from results without renumbering other placements.
    out: List[Entry] = []
    for e in entries:
        if any(_is_tony_spredeman(p) for p in e.players):
            continue
        out.append(e)
    return out


def _find_all_places(obj: object) -> List[Entry]:
    """Heuristic: search the Next.js payload for objects that look like standings rows."""
    entries: List[Entry] = []

    def walk(x: object) -> None:
        if isinstance(x, dict):
            # Common keys: place/rank/position; players/team
            if ("place" in x or "rank" in x or "position" in x) and ("players" in x or "team" in x or "competitor" in x):
                place_val = x.get("place", x.get("rank", x.get("position")))
                try:
                    place = int(place_val)
                except Exception:
                    place = None

                players: List[str] = []
                if "players" in x and isinstance(x["players"], list):
                    for p in x["players"]:
                        if isinstance(p, dict):
                            n = p.get("name") or p.get("displayName")
                            if isinstance(n, str):
                                players.append(n)
                        elif isinstance(p, str):
                            players.append(p)
                elif "team" in x and isinstance(x["team"], dict):
                    members = x["team"].get("players") or x["team"].get("members")
                    if isinstance(members, list):
                        for p in members:
                            if isinstance(p, dict):
                                n = p.get("name") or p.get("displayName")
                                if isinstance(n, str):
                                    players.append(n)
                elif "competitor" in x and isinstance(x["competitor"], dict):
                    n = x["competitor"].get("name") or x["competitor"].get("displayName")
                    if isinstance(n, str):
                        players.append(n)

                if place is not None and players:
                    players = [pretty_name(p.strip()) for p in players if p and p.strip()]
                    if players:
                        entries.append(Entry(place=place, players=players))

            for v in x.values():
                walk(v)
        elif isinstance(x, list):
            for v in x:
                walk(v)

    walk(obj)

    # Deduplicate by (place, players)
    seen: Set[Tuple[int, Tuple[str, ...]]] = set()
    out: List[Entry] = []
    for e in entries:
        key = (e.place, tuple(e.players))
        if key not in seen:
            seen.add(key)
            out.append(e)

    # Sort by place; stable-ish
    out.sort(key=lambda e: (e.place, "|".join(e.players)))
    return out


def _render_entries(entries: Sequence[Entry]) -> List[str]:
    lines: List[str] = []
    prev_place: Optional[int] = None
    for e in entries:
        payload = "|".join(e.players)
        if prev_place is None or e.place != prev_place:
            lines.append(f"{e.place}.{payload}")
        else:
            lines.append(payload)
        prev_place = e.place
    return lines


def _all_players_match(open_entries: Sequence[Entry], pred) -> bool:
    for e in open_entries:
        for p in e.players:
            if not pred(p):
                return False
    return True


def _rerank_individuals_by_open_place(open_entries: Sequence[Entry], pred) -> List[str]:
    """For open events, produce a ranked list of *individuals* matching pred.

    - Preserves ties: same open place => same rank in the subgroup.
    - A doubles team with 1 woman contributes 1 entry (just the woman).
    - A doubles team with 2 women contributes 2 entries (both women) tied at that open place.
    """
    picked: List[Tuple[int, str]] = []
    for e in open_entries:
        for p in e.players:
            if pred(p):
                picked.append((e.place, p))

    out: List[str] = []
    last_open_place: Optional[int] = None
    rank = 0
    for open_place, name in picked:
        if last_open_place is None or open_place != last_open_place:
            rank += 1
            out.append(f"{rank}.{name}")
        else:
            out.append(name)
        last_open_place = open_place
    return out


def _club_from_url(url: str) -> str:
    m = TOURNAMENT_URL_RE.match(url)
    if not m:
        raise ValueError(f"Unrecognized Kickertool URL: {url}")
    return m.group("club")


def _tournament_id_from_url(url: str) -> str:
    m = TOURNAMENT_URL_RE.match(url)
    if not m:
        raise ValueError(f"Unrecognized Kickertool URL: {url}")
    return m.group("tid")


def _tournament_title(next_data: dict) -> str:
    v = next_data.get("name")
    if isinstance(v, str) and v.strip():
        return v.strip()
    return "Unknown Tournament"


def safe_folder_name(name: str) -> str:
    name = name.strip()
    name = re.sub(r"[\\/:*?\"<>|]", "-", name)
    name = re.sub(r"\s+", " ", name)
    return name


def write_lines(path: str, lines: List[str], overwrite: bool) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if (not overwrite) and os.path.exists(path):
        return
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + ("\n" if lines else ""))


def delete_if_exists(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        return


def main() -> None:
    ap = argparse.ArgumentParser(description="Extract Kickertool standings and write FRFM .txt files")
    ap.add_argument("--root", default=".", help="Repo root (default: .)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing .txt files")
    ap.add_argument("--timeout", type=int, default=30)
    ap.add_argument(
        "--excel",
        default="",
        help="Path to rankings Excel used as source of truth for women (sheets containing 'women')",
    )
    ap.add_argument(
        "--juniors",
        default="",
        help="Comma-separated list of junior player names (only these will be classified as juniors)",
    )
    ap.add_argument("urls", nargs="+", help="Kickertool standings URLs")
    args = ap.parse_args()

    if not args.excel.strip():
        raise ValueError("--excel is required (women allowlist comes from Excel)")

    excel_path = args.excel
    if not excel_path.lower().endswith(".xlsx"):
        excel_path = excel_path + ".xlsx"
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(args.root, excel_path)

    women_allow = load_women_allowlist_from_excel(excel_path)

    junior_allow: Set[str] = set()
    if args.juniors.strip():
        for raw in args.juniors.split(","):
            n = raw.strip()
            if not n:
                continue
            nn = _norm(n)
            junior_allow.add(nn)
            swapped = _swap_two_tokens(nn)
            if swapped:
                junior_allow.add(swapped)
    else:
        raise ValueError("--juniors is required (only the provided juniors are categorized as juniors)")

    def is_woman(name: str) -> bool:
        return _norm(name) in women_allow

    def is_junior(name: str) -> bool:
        return _norm(name) in junior_allow

    extracted: List[Tuple[str, str]] = []  # (club, tournament_dir)

    for url in args.urls:
        club = _club_from_url(url)
        tid = _tournament_id_from_url(url)
        tj = fetch_tournament_json(tid, timeout_s=args.timeout)
        title = _tournament_title(tj)
        tdir = os.path.join(args.root, club, safe_folder_name(title))

        title_lc = title.lower()
        women_only_event = "women" in title_lc
        junior_only_event = "junior" in title_lc

        open_entries = remove_tony(extract_final_standings(tj))

        mens_lines = _render_entries(open_entries)
        write_lines(os.path.join(tdir, "mens.txt"), mens_lines, overwrite=args.overwrite)

        # If the whole tournament is already women-only (e.g. Women Doubles), keep team entries.
        # Otherwise, in open events extract only the women as individuals.
        womens_lines = (
            _render_entries(open_entries)
            if women_only_event or _all_players_match(open_entries, is_woman)
            else _rerank_individuals_by_open_place(open_entries, is_woman)
        )
        womens_path = os.path.join(tdir, "womens.txt")
        if womens_lines:
            write_lines(womens_path, womens_lines, overwrite=args.overwrite)
        elif args.overwrite:
            delete_if_exists(womens_path)

        junior_lines = (
            _render_entries(open_entries)
            if junior_only_event or _all_players_match(open_entries, is_junior)
            else _rerank_individuals_by_open_place(open_entries, is_junior)
        )
        junior_path = os.path.join(tdir, "junior.txt")
        if junior_lines:
            write_lines(junior_path, junior_lines, overwrite=args.overwrite)
        elif args.overwrite:
            delete_if_exists(junior_path)

        extracted.append((club, tdir))
        print(f"Wrote: {tdir}")

    print(f"Done. Extracted {len(extracted)} tournaments.")


if __name__ == "__main__":
    main()
